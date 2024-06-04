from django.shortcuts import render, HttpResponseRedirect
from django.contrib.auth import logout
from django.http import JsonResponse, HttpResponse
from .models import GroupsTable, SubgroupsTable, AuthUser, StudentsTable, AttendanceTable, WeeksTable, CommentsTable, AccessWeeksTable, AuthUserGroups
from django.contrib.auth.decorators import login_required
from webway.decorators import group_required
from django.contrib.auth.hashers import make_password
from datetime import datetime
from django.db.models import Q
import json
from django.urls import reverse
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import os
import re
from django.core.mail import send_mail
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt

base_path = "C:\\archive\\"

@login_required
@group_required('admin')
def index(request):
    success = request.GET.get('success', 'false')
    groups = GroupsTable.objects.all()
    return render(request, 'adminpages/startpage.html', {'groups': groups, 'success': success})

@login_required
@group_required('admin')
def feedback(request):
    return render(request, 'adminpages/feedback.html')

@login_required
@group_required('admin')
def send_feedback(request):
    if request.method == 'POST':
        feedback_text = request.POST.get('feedbackText')
        send_mail(
            'Отзыв от пользователя',
            feedback_text,
            'PPP@mail.ru',
            ['PPP@mail.ru'],
            fail_silently=True,
        )
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})

@login_required
@group_required('admin')
def subgroups_select(request, key_group):
    subgroups = SubgroupsTable.objects.filter(key_group=key_group)
    group = GroupsTable.objects.get(id_groups_table=key_group)
    success = request.GET.get('success', 'false')
    return render(request, 'adminpages/subgroups_select.html', {'subgroups': subgroups, 'group': group, 'success': success})

@login_required
@group_required('admin')
def table_students(request, key_subgroup):
    students = StudentsTable.objects.filter(key_subgroup=key_subgroup).order_by('student_name')
    subgroup = SubgroupsTable.objects.get(id_subgroups_table=key_subgroup)
    weeks = WeeksTable.objects.all().order_by('id_weeks_table')
    attendance_data = []
    AccessWeeks = AccessWeeksTable.objects.filter(subgroup=key_subgroup).order_by('week')

    for student in students:
            student_attendance = {}
            attendance_records = AttendanceTable.objects.filter(id_student=student.id_students_table)
            comments = CommentsTable.objects.filter(id_student=student.id_students_table)
            for week in weeks:
                student_attendance[week.id_weeks_table] = {'attendance_disrespectfully': None,'attendance_respectfully': None,'comment': ''}
            for record in attendance_records:
                student_attendance[record.id_week.id_weeks_table].update({'attendance_disrespectfully': record.attendance_disrespectfully,'attendance_respectfully': record.attendance_respectfully})
            for comment in comments:
                student_attendance[comment.id_week.id_weeks_table].update({'comment': '' if comment.comment == 'None' else comment.comment,'color': comment.color})
            attendance_data.append({'student': student, 'attendance': student_attendance})
    month_data = []
    prev_month = None
    month_span = 0
    for week in weeks:
        start_date_str, end_date_str = week.week_name.split(' - ')
        start_date = datetime.strptime(start_date_str, '%d.%m')
        month_name = {
            1: 'Январь',
            2: 'Февраль',
            3: 'Март',
            4: 'Апрель',
            5: 'Май',
            6: 'Июнь',
            7: 'Июль',
            8: 'Август',
            9: 'Сентябрь',
            10: 'Октябрь',
            11: 'Ноябрь',
            12: 'Декабрь'
        }[start_date.month]

        if month_name != prev_month:
            if prev_month is not None:
                month_data.append({
                    'month_name': prev_month,
                    'month_span': 2 * month_span
                })
            month_span = 0
            prev_month = month_name

        month_span += 1

    if prev_month is not None:
        month_data.append({
            'month_name': prev_month,
            'month_span': 2 * month_span
        })
        
    access_weeks = AccessWeeksTable.objects.filter(subgroup=subgroup)
    for week in weeks:
        access_week = access_weeks.filter(week_id=week.id_weeks_table, color_sent='Y').first()
        if access_week:
            week.color = 'background: yellow;'
        else:
            week.color = ''

    return render(request, 'adminpages/table_students.html', {'attendance_data': attendance_data, 'subgroup': subgroup, 'weeks': weeks, 'month_data': month_data, 'AccessWeeks': AccessWeeks})

@login_required
@group_required('admin')
def logout_view(request):
    logout(request)
    return HttpResponseRedirect('/')

@login_required
@group_required('admin')
def add_lock(request, n):
    if request.method == 'POST':
        week = WeeksTable.objects.get(id_weeks_table = request.POST.get('week'))
        key_subgroup_id = SubgroupsTable.objects.get(id_subgroups_table = request.POST.get('key_subgroup'))   
        access_week_entry, created = AccessWeeksTable.objects.get_or_create(subgroup = key_subgroup_id, week = week)
        if not created:
            AccessWeeksTable.objects.filter(week=week, subgroup=key_subgroup_id).delete()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})
    
@login_required
@group_required('admin')
def save_table(request, n):
    if request.method == 'POST':
        try:
            data_attendance = json.loads(request.POST['dataAttendance'])
            data_comments = json.loads(request.POST['dataComments'])
            for item in data_attendance:
                week = item['week_id']
                id_student = item['student_id']
                respectfully = item['attendance_respectfully']
                disrespectfully = item['attendance_disrespectfully']
                if (respectfully == '0' or respectfully == ''):
                    respectfully = None
                if (disrespectfully == '0' or disrespectfully == ''):
                    disrespectfully = None
                if (respectfully == disrespectfully == None):
                    AttendanceTable.objects.filter(id_student_id=id_student, id_week_id=week).delete()
                else:
                    attendance = AttendanceTable.objects.filter(id_student=id_student, id_week=week).first()
                    if attendance:
                        AttendanceTable.objects.filter(id_student=id_student, id_week=week).update(attendance_respectfully=respectfully,attendance_disrespectfully=disrespectfully)
                    else:
                        attendance = AttendanceTable.objects.create(
                            id_student_id=id_student,
                            id_week_id=week,
                            attendance_respectfully=respectfully,
                            attendance_disrespectfully=disrespectfully
                        )

            for item in data_comments:
                week = item['week_id']
                id_student = item['student_id']
                comment = item['comment']
                color = item['color']
                if (comment == '' or comment == 'None'):
                    comment = None
                if (color == 'rgba(0, 0, 0, 0)' or color == 'rgba(255, 255, 255, 0.5)'):
                    color = None
                if (comment == color == None):
                    CommentsTable.objects.filter(id_student_id=id_student, id_week_id=week).delete()
                else:
                    search = CommentsTable.objects.filter(id_student=id_student, id_week=week).first()
                    if search:
                        CommentsTable.objects.filter(id_student=id_student, id_week=week).update(comment=comment,color=color)
                    else:
                        comment = CommentsTable.objects.create(
                            id_student_id=id_student,
                            id_week_id=week,
                            comment=comment,
                            color=color
                        )
            subgroup = request.POST['subgroup']
            AccessWeeksTable.objects.filter(subgroup=subgroup, color_sent='Y').delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})

@login_required
@group_required('admin')
def search_students(request):
    if request.method == 'POST':
        searchName = request.POST.get('SearchStudentName')
        searchCol = int(request.POST.get('SearchAttendance'))
        result = []
        if searchName:
            searchStudents = StudentsTable.objects.filter(Q(student_name__icontains=searchName))
        else:
            searchStudents = StudentsTable.objects.all()
        for student in searchStudents:     
            searchAttendances = AttendanceTable.objects.filter(id_student = student.id_students_table)
            sumAttendance = 0
            for attendance in searchAttendances:
                sumAttendance += (attendance.attendance_respectfully or 0) + (attendance.attendance_disrespectfully or 0)
            if sumAttendance >= searchCol:
                result.append({'subgroupName': student.key_subgroup.subgroup_name, 'subgroupId': reverse('table_students_admin', args=[student.key_subgroup.id_subgroups_table]), 'name': student.student_name, 'attendances': sumAttendance})
        return JsonResponse({'success': True, 'result': result})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})

@login_required
@group_required('admin')
def addGroup(request):
    if request.method == 'POST':
        nameGroup = request.POST['groupName']
        Elsegroup = GroupsTable.objects.filter(groups_name = nameGroup)
        if Elsegroup.exists():
            return JsonResponse({'success': False, 'message': 'Направление с таким названием уже существует'})
        new_group = GroupsTable.objects.create(groups_name = nameGroup)
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})
    
@login_required
@group_required('admin')
def delGroup(request):
    if request.method == 'POST':
        idGroup = request.POST['groupId']
        group_to_delete = GroupsTable.objects.filter(id_groups_table = idGroup)
        group_to_delete.delete()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})
    
@login_required
@group_required('admin')
def saveGroup(request):
    if request.method == 'POST':
        nameGroup = request.POST['groupName']
        idGroup = request.POST['groupId']
        Elsegroup = GroupsTable.objects.filter(groups_name = nameGroup)
        if Elsegroup.exists():
            return JsonResponse({'success': False, 'message': 'Направление с таким названием уже существует'})
        group_to_update = GroupsTable.objects.get(id_groups_table = idGroup)
        group_to_update.groups_name = nameGroup
        group_to_update.save()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})

@login_required
@group_required('admin')
def accounts_admin(request):
    groups = GroupsTable.objects.all()
    return render(request, 'adminpages/accounts_admin.html', {'groups': groups})

@login_required
@group_required('admin')
def load_subgroups(request):
    group_table_id = request.GET.get('group_table_id')
    if group_table_id != 'Null':
        subgroups = SubgroupsTable.objects.filter(key_group=group_table_id).all()
        return JsonResponse(list(subgroups.values('id_subgroups_table', 'subgroup_name')), safe=False)
    else:
        return JsonResponse([{'id_subgroups_table': 0, 'subgroup_name': 'Без группы'}], safe=False)

@login_required
@group_required('admin')
def create_account(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        hashed_password = make_password(password)
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        patronymic = request.POST.get('patronymic', '')
        email = request.POST['email']
        access_level = request.POST['access_level']
        subgroup_id = request.POST.get('subgroup', None)
        Dsubgroup = None
        if access_level == '2' and (subgroup_id == None or subgroup_id == ''):
            return JsonResponse({'success': False, 'message': 'Невозможно создать старосту группы без группы!'})
        if subgroup_id:
            Dsubgroup = SubgroupsTable.objects.get(id_subgroups_table = subgroup_id)
        try:
            Duser = AuthUser.objects.filter(username=username).first()
            if not Duser:
                user = AuthUser.objects.create(
                    username=username,
                    password=hashed_password,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    patronymic=patronymic,
                    key_subgroups=Dsubgroup,
                    is_superuser = 0,
                    is_staff = 0,
                    is_active = 1,
                    date_joined = timezone.now()
                )
                user.save()

                user_group = AuthUserGroups.objects.create(
                    user_id=user.id,
                    group_id=access_level
                )
                
                return JsonResponse({'success': True, 'message': 'Аккаунт успешно создан'})
            else:
                return JsonResponse({'success': False, 'message': 'Пользователь с таким логином уже существует!'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})

@login_required
@group_required('admin')
def ajax_search_accounts(request):
    access_level = request.GET.get('access_level')
    username = request.GET.get('username')
    first_name = request.GET.get('first_name')
    last_name = request.GET.get('last_name')
    patronymic = request.GET.get('patronymic')
    group_id = request.GET.get('group_id')
    subgroup_id = request.GET.get('subgroup_id')
    
    users = AuthUser.objects.all()

    if access_level:
        users = users.filter(authusergroups__group_id=access_level)
    if username:
        users = users.filter(username__icontains=username)
    if first_name:
        users = users.filter(first_name__icontains=first_name)
    if last_name:
        users = users.filter(last_name__icontains=last_name)
    if patronymic:
        users = users.filter(patronymic__icontains=patronymic)
    if group_id and group_id != 'Null':
        users = users.filter(key_subgroups__key_group=group_id)
    if subgroup_id:
        users = users.filter(key_subgroups=subgroup_id)

    results = []

    if access_level == '2':
        for user in users:
            subgroup = user.key_subgroups
            group = subgroup.key_group
            results.append({
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'patronymic': user.patronymic,
                'group': {'id': group.id_groups_table, 'name': group.groups_name},
                'subgroup': {'id': subgroup.id_subgroups_table, 'name': subgroup.subgroup_name},
            })
    else:
        for user in users:
            results.append({
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'patronymic': user.patronymic,
                'group': {'id': '0', 'name': 'Нет направления'},
                'subgroup': {'id': '0', 'name': 'Нет группы'},
            })
    return JsonResponse({'results': results})

@login_required
@group_required('admin')
@csrf_exempt
def ajax_delete_account(request):
    if request.method == 'POST':
        account_id = request.POST.get('account_id')
        try:
            user = AuthUser.objects.get(id=account_id)
            usergroup = AuthUserGroups.objects.get(user_id = account_id)
            usergroup.delete()
            user.delete()
            return JsonResponse({'success': True, 'message': 'Аккаунт успешно удален'})
        except AuthUser.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Аккаунт не найден'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})

@login_required
@group_required('admin')
@csrf_exempt
def ajax_save_account(request):
    if request.method == 'POST':
        account_id = request.POST.get('account_id')
        group_id = request.POST.get('group_id')
        subgroup_id = request.POST.get('subgroup_id')
        access_level = request.POST.get('access_level')
        try:
            user = AuthUser.objects.get(id=account_id)
            if access_level == '2':
                if group_id and group_id != 'Null':
                    user.key_subgroups.group_id = GroupsTable.objects.get(id_groups_table=group_id)
                else:
                    return JsonResponse({'success': False, 'message': 'Не выбрано направление!'})
                if subgroup_id:
                    user.key_subgroups = SubgroupsTable.objects.get(id_subgroups_table=subgroup_id)
                else:
                    return JsonResponse({'success': False, 'message': 'Не выбрана группа!'})
                user.save()
            return JsonResponse({'success': True, 'message': 'Аккаунт успешно обновлен'})
        except AuthUser.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Аккаунт не найден'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})

@login_required
@group_required('admin')
def groups_admin(request):
    groups = GroupsTable.objects.all().order_by('groups_name')
    return render(request, 'adminpages/groups_admin.html', {'groups': groups})

@login_required
@group_required('admin')
def subgroups_admin(request, key_group):
    group = GroupsTable.objects.get(id_groups_table = key_group)
    subgroups = SubgroupsTable.objects.filter(key_group = key_group).order_by('subgroup_name')
    return render(request, 'adminpages/subgroups_admin.html', {'group': group, 'subgroups': subgroups})

@login_required
@group_required('admin')
def addSubGroup_admin(request, n):
    if request.method == 'POST':
        nameSubGroup = request.POST['subgroupName']
        id_group = request.POST['key_group']
        group = GroupsTable.objects.get(id_groups_table = id_group)
        Elsesubgroup = SubgroupsTable.objects.filter(subgroup_name = nameSubGroup)
        if Elsesubgroup.exists():
            return JsonResponse({'success': False, 'message': 'Группа с таким названием уже существует'})
        new_subgroup = SubgroupsTable.objects.create(subgroup_name = nameSubGroup, key_group = group)

        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})
    
@login_required
@group_required('admin')
def delSubGroup(request, n):
    if request.method == 'POST':
        idSubGroup = request.POST['subgroupId']
        subgroup_to_delete = SubgroupsTable.objects.filter(id_subgroups_table = idSubGroup)
        subgroup_to_delete.delete()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})
    
@login_required
@group_required('admin')
def saveSubGroup(request, n):
    if request.method == 'POST':
        nameSubGroup = request.POST['subgroupName']
        idSubGroup = request.POST['subgroupId']
        Elsesubgroup = SubgroupsTable.objects.filter(subgroup_name = nameSubGroup)
        if Elsesubgroup.exists():
            return JsonResponse({'success': False, 'message': 'Группа с таким названием уже существует'})
        subgroup_to_update = SubgroupsTable.objects.get(id_subgroups_table = idSubGroup)
        subgroup_to_update.subgroup_name = nameSubGroup
        subgroup_to_update.save()
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})

@login_required
@group_required('admin')
def archive(request):
    try:
        academic_years = [d for d in os.listdir(base_path) if os.path.isdir(os.path.join(base_path, d))]

        def extract_years(folder_name):
            match = re.match(r'(\d{4})-(\d{4}) учебный год', folder_name)
            if match:
                start_year = int(match.group(1))
                end_year = int(match.group(2))
                return start_year, end_year
            return None
        academic_years.sort(key=lambda x: extract_years(x) if extract_years(x) else (0, 0), reverse=True)

    except FileNotFoundError:
        academic_years = []
    return render(request, 'adminpages/archive.html', {'academic_years': academic_years, 'base_path': base_path})

@login_required
@group_required('admin')
def archive_subgroups(request, year, semester):
    path = os.path.join(base_path, year, semester)
    try:
        subgroups_with_extension = [f for f in os.listdir(path) if os.path.isfile(os.path.join(path, f)) and f.endswith('.xlsx')]
        subgroups = [os.path.splitext(f)[0] for f in subgroups_with_extension]
    except FileNotFoundError:
        subgroups = []
    return render(request, 'adminpages/archive_subgroups.html', {'year': year, 'semester': semester, 'subgroups': subgroups, 'base_path': base_path})

@login_required
@group_required('admin')
def download_excel(request, year, semester, subgroup):
    excel_file_path = os.path.join(base_path, year, semester, f"{subgroup}.xlsx")
    if os.path.exists(excel_file_path):
        with open(excel_file_path, 'rb') as excel_file:
            file_content = excel_file.read()
        
        response = HttpResponse(file_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{subgroup}.xlsx"'
        return response
    else:
        return HttpResponse(status=404)

@login_required
@group_required('admin')
def archiving(request):
    current_year = datetime.now().year

    weeks = WeeksTable.objects.all().order_by('id_weeks_table')
    first_week_start_date_str = weeks[0].week_name.split(' - ')[0]
    first_week_start_date = datetime.strptime(first_week_start_date_str, '%d.%m')
    first_week_month = first_week_start_date.month
    
    if first_week_month in [9, 10, 11, 12]:
        semester = "I семестр"
        current_year -= 1
    else:
        semester = "II семестр"
    return render(request, 'adminpages/archiving.html', {'semester': semester, 'year': current_year})

def rgba_to_hex(rgba):
    if rgba:
        rgba_list = rgba.split("(")[1].split(")")[0].split(",")
        r, g, b = [int(x) for x in rgba_list[:3]]
        hex_color = "{:02x}{:02x}{:02x}".format(r, g, b)
        return hex_color
    else:
        return None
    
def getMonths(weeks):
    month_data = []
    dataForMonths = []
    prev_month = None
    month_span = 0
    for week in weeks:
        start_date_str, end_date_str = week.week_name.split(' - ')
        start_date = datetime.strptime(start_date_str, '%d.%m')
        month_name = {
            1: 'Январь',
            2: 'Февраль',
            3: 'Март',
            4: 'Апрель',
            5: 'Май',
            6: 'Июнь',
            7: 'Июль',
            8: 'Август',
            9: 'Сентябрь',
            10: 'Октябрь',
            11: 'Ноябрь',
            12: 'Декабрь'
        }[start_date.month]

        if month_name != prev_month:
            if prev_month is not None:
                dataForMonths.append(month_span)
                month_data.append(prev_month)
                for i in range(month_span - 1):
                    month_data.append('')
            month_span = 0
            prev_month = month_name

        month_span += 1

    if prev_month is not None:
        dataForMonths.append(month_span)
        month_data.append(prev_month)
        for i in range(month_span - 1):
            month_data.append('')
    return month_data, dataForMonths

@login_required
@group_required('admin')
def archivingPost(request):
    if request.method == 'POST':
        try:
            
            font = Font(size=16)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            borderRight = Border(left=Side(style='thin'), right=Side(style='thick'), top=Side(style='thin'), bottom=Side(style='thin'))
            borderLeft = Border(left=Side(style='thick'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            weeksNew = json.loads(request.POST['weeksNew'])
            year = request.POST['year']
            semester = request.POST['semester']
            if semester == 'I семестр':
                year = str(year) + '-' + str(int(year) + 1)
            else:
                year = str(int(year) - 1) + '-' + str(year)

            
            os.makedirs(base_path, exist_ok=True)
            year_path = os.path.join(base_path, year + ' учебный год')
            os.makedirs(year_path, exist_ok=True)
            semester_path = os.path.join(year_path, semester)
            os.makedirs(semester_path, exist_ok=True)

            groups = GroupsTable.objects.all().order_by('groups_name')
            weeks = WeeksTable.objects.all().order_by('id_weeks_table')
            rowFirst = []
            for week in weeks:
                start_date, end_date = map(str.strip, week.week_name.split('-'))
                start_day = start_date[:2]
                end_day = end_date[:2]
                rowFirst.append([start_day + '-' + end_day, '', ''])
            months, dataMonths = getMonths(weeks)
            rowSecond = []        
            for month in months:
                rowSecond.append([month, '', ''])
            for group in groups:
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)
                subgroups = SubgroupsTable.objects.filter(key_group=group.id_groups_table).order_by('subgroup_name')
                if subgroups:
                    for subgroup in subgroups:
                        sheet = workbook.create_sheet(title=subgroup.subgroup_name)
                        studentsInSubgroup = StudentsTable.objects.filter(key_subgroup=subgroup.id_subgroups_table).order_by('student_name')
                        data = [[['По уважительной / По неуважительной','',''], *rowFirst]]
                        data.append([[semester + ': ' + year + '  ' + subgroup.subgroup_name, '', ''], *rowSecond])
                        if studentsInSubgroup:
                            for student in studentsInSubgroup:
                                cellsOneStudent = [[student.student_name, '', '']]
                                for week in weeks:
                                    attendance_item = AttendanceTable.objects.filter(id_student=student.id_students_table, id_week=week.id_weeks_table).first()
                                    comment_item = CommentsTable.objects.filter(id_student=student.id_students_table, id_week=week.id_weeks_table).first()
                                    cell_data = []
                                    if attendance_item:
                                        cell_data.append(str(attendance_item.attendance_respectfully if attendance_item.attendance_respectfully is not None else '') + '/' + str(attendance_item.attendance_disrespectfully if attendance_item.attendance_disrespectfully is not None else ''))
                                    else:
                                        cell_data.append('')
                                    if comment_item:
                                        cell_data.append(comment_item.comment if comment_item.comment is not None else '')
                                        cell_data.append(comment_item.color if comment_item.color is not None else '')
                                    else:
                                        cell_data.extend(['', ''])
                                    cellsOneStudent.append(cell_data)
                                data.append(cellsOneStudent)
                        for col in range(len(data[0])):
                            column_letter = openpyxl.utils.get_column_letter(col + 1)
                            sheet.column_dimensions[column_letter].auto_size = True
                            if col == 0:
                                sheet.column_dimensions[column_letter].width = 52
                            else:
                                sheet.column_dimensions[column_letter].width = 18
                            for row in range(len(data)):
                                cell_value, comment, color = data[row][col]
                                cell = sheet.cell(row=row + 1, column=col + 1)
                                cell.value = cell_value
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.font = font
                                cell.border = border
                                if comment:
                                    cell.comment = Comment(comment, "Author")
                                if color:
                                    hex_color = rgba_to_hex(color)
                                    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                                    cell.fill = fill
                                if row == 1:
                                    fill = PatternFill(start_color='da9694', end_color='da9694', fill_type="solid")
                                    cell.fill = fill
                        merge_start = 2
                        for merge in dataMonths:
                            endColumn = merge+merge_start-1
                            sheet.merge_cells(start_row=2, start_column=merge_start, end_row=2, end_column=endColumn)
                            for row in range(len(data)):
                                sheet.cell(row=row + 1, column=merge_start).border = borderLeft
                                sheet.cell(row=row + 1, column=endColumn).border = borderRight
                            merge_start = endColumn + 1
                    for sheet in workbook:
                        sheet.protection.sheet = True
                        sheet.protection.password = "На горшке сидел король"
                    file_path = os.path.join(semester_path, group.groups_name + ".xlsx")
                    workbook.save(file_path)
            CommentsTable.objects.all().delete()
            AttendanceTable.objects.all().delete()
            WeeksTable.objects.all().delete()
            for week in weeksNew:
                WeeksTable.objects.create(week_name = week)
            Dsubgroups = SubgroupsTable.objects.all()
            Dweeks = WeeksTable.objects.all()
            for week in Dweeks:
                for subgroup in Dsubgroups:
                    AccessWeeksTable.objects.create(week = week, subgroup = subgroup)

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    else:
        return JsonResponse({'success': False, 'message': 'Метод запроса не поддерживается'})